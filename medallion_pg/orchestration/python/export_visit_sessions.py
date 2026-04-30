"""Daily visit-sessions report exporter.

Pulls one row per (visit x overlapping game session) from gold.fn_visit_sessions,
writes an .xlsx matching the legacy Casino Daily Analysis template, and emails it.

Recipient routing:
  - REPORT_TO_EMAILS_SAFE: always-on recipients (default: data@)
  - REPORT_TO_EMAILS_TEAM: added only when the data pipeline is healthy
    (i.e., gold.fact_membership_day has rows for the report date)

If the pipeline is degraded (no rows for the target gaming day), the team list
is dropped and the email goes only to SAFE recipients with a softer "data may
be incomplete" subject and body. The team is intentionally not paged about
broken runs.

SMTP profile: by default reuses the global SMTP_* settings. Set REPORT_SMTP_*
env vars to use a separate mail server for the report only (so the report can
send From: data@tourinvestgroup.com without affecting the pipeline-status
email's relay).

Run standalone (Task Scheduler at 07:00) — no Prefect dependency.
Defaults to yesterday (UTC).
"""
from __future__ import annotations

import argparse
import os
import smtplib
import sys
from datetime import date, datetime, time as dtime, timedelta, timezone
from email.message import EmailMessage
from pathlib import Path

import psycopg2
from dotenv import load_dotenv
from openpyxl import Workbook

BASE_DIR = Path(__file__).resolve().parent
PROJECT_DIR = BASE_DIR.parent.parent
REPORTS_DIR = PROJECT_DIR / "reports"

load_dotenv(BASE_DIR / ".env")

HEADERS = [
    "GamingDay",
    "Membership",
    "VisitNo",
    "CasinoEntryTime",
    "CasinoExitTime",
    "SessionStart",
    "SessionFinish",
    "averBet",
]


def parse_args() -> argparse.Namespace:
    yesterday = (datetime.now(timezone.utc).date() - timedelta(days=1)).isoformat()
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--from-date", default=yesterday, help="YYYY-MM-DD (default: yesterday UTC)")
    parser.add_argument("--to-date", default=yesterday, help="YYYY-MM-DD (default: yesterday UTC)")
    parser.add_argument("--no-email", action="store_true", help="Generate the file only, skip SMTP")
    parser.add_argument("--force-degraded", action="store_true",
                        help="Force degraded mode (skip team list) — for testing the safety path")
    return parser.parse_args()


def fetch_rows(pg_conn_str: str, from_date: str, to_date: str) -> list[tuple]:
    with psycopg2.connect(pg_conn_str) as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT gaming_day, membership, visit_no, casino_entry_time, casino_exit_time,
                       session_start, session_finish, averbet
                FROM gold.fn_visit_sessions(%s::date, %s::date)
                """,
                (from_date, to_date),
            )
            return cur.fetchall()


def pipeline_is_healthy(pg_conn_str: str, target_date: str) -> bool:
    """Return True iff gold.fact_membership_day has any rows for target_date.

    Healthy means the daily DW pipeline has produced data for the report date.
    Degraded means the pipeline didn't run, failed, or hadn't reached this date
    yet — in which case the report goes only to the SAFE recipient list.
    """
    with psycopg2.connect(pg_conn_str) as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT COUNT(*) FROM gold.fact_membership_day WHERE gamingday = %s::date",
                (target_date,),
            )
            return cur.fetchone()[0] > 0


def resolve_recipients(healthy: bool) -> tuple[list[str], list[str]]:
    """Return (recipients, dropped_team_recipients) given pipeline health."""
    safe = [e.strip() for e in (os.getenv("REPORT_TO_EMAILS_SAFE") or "").split(",") if e.strip()]
    team = [e.strip() for e in (os.getenv("REPORT_TO_EMAILS_TEAM") or "").split(",") if e.strip()]
    if not safe:
        raise RuntimeError("REPORT_TO_EMAILS_SAFE must be set (at minimum)")
    if healthy:
        return safe + team, []
    return safe, team


def _trunc_minute(t):
    """Drop seconds + microseconds. Matches the legacy MSSQL `CONVERT(char(5), ..., 108)`
    which output strings like '16:07'. SSMS-exported xlsx stored those as time(h,m,0)."""
    if t is None:
        return None
    return t.replace(second=0, microsecond=0)


def write_xlsx(rows: list[tuple], out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(HEADERS)

    for gd, mem, vn, in_t, out_t, ss, sf, ab in rows:
        ws.append([
            datetime.combine(gd, dtime.min) if gd is not None else None,
            mem,
            vn,
            _trunc_minute(in_t),
            _trunc_minute(out_t),
            _trunc_minute(ss) if ss is not None else "NULL",
            _trunc_minute(sf) if sf is not None else "NULL",
            float(ab) if ab is not None else "NULL",
        ])

    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=1).number_format = "mm-dd-yy"
        ws.cell(row=r, column=4).number_format = "h:mm"
        ws.cell(row=r, column=5).number_format = "h:mm"
        for col in (6, 7):
            cell = ws.cell(row=r, column=col)
            if cell.value != "NULL":
                cell.number_format = "h:mm"

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def _friendly_date(d: date) -> str:
    """e.g. date(2026, 4, 8) -> 'Wed, Apr 8, 2026' (no leading zero on day)."""
    return f"{d.strftime('%a, %b')} {d.day}, {d.year}"


def _build_subject_and_body(
    from_date: str,
    to_date: str,
    row_count: int,
    healthy: bool,
) -> tuple[str, str]:
    """Construct (subject, body) in the team-friendly tone.

    Healthy mode: warm + brief greeting, sign-off as 'Tourinvest Data Team'.
    Degraded mode: softer alert, ⚠ subject marker, plain-language warning."""
    from_d = date.fromisoformat(from_date)
    to_d = date.fromisoformat(to_date)
    yesterday = datetime.now(timezone.utc).date() - timedelta(days=1)

    if from_d == to_d:
        date_label = _friendly_date(to_d)
        short = f"{to_d.strftime('%b')} {to_d.day}"
    else:
        date_label = f"{_friendly_date(from_d)} – {_friendly_date(to_d)}"
        short = date_label

    if healthy:
        subject = f"Daily Visits Report — {date_label}"
        if from_d == to_d == yesterday:
            line = f"Attached is yesterday's casino visits report ({row_count} entries from {short})."
        elif from_d == to_d:
            line = f"Attached is the casino visits report for {short} ({row_count} entries)."
        else:
            line = f"Attached is the casino visits report for {short} ({row_count} entries)."
        body = "\n".join([
            "Hi team,",
            "",
            line,
            "",
            "Questions or feedback? Just reply — it'll reach the data team.",
            "",
            "— Tourinvest Data Team",
        ]) + "\n"
    else:
        subject = f"⚠ Daily Visits Report — {date_label} (data may be incomplete)"
        body = "\n".join([
            "Hi,",
            "",
            "Yesterday's report is attached, but heads up: today's data load didn't",
            "complete on time, so the file may be empty or incomplete.",
            "",
            "The wider team has not been emailed for this run. Once data is back,",
            "we'll re-send the report.",
            "",
            "— Tourinvest Data Team",
        ]) + "\n"

    return subject, body


def _resolve_smtp_config() -> dict:
    """Return SMTP config dict.

    REPORT_SMTP_* values override the global SMTP_* values per-field, so the
    report can use a separate mail server while the pipeline-status email
    keeps its existing relay. Any unset REPORT_SMTP_* falls back to SMTP_*."""
    def pick(report_key: str, global_key: str, default: str | None = None) -> str | None:
        return os.getenv(report_key) or os.getenv(global_key) or default

    host = pick("REPORT_SMTP_HOST", "SMTP_HOST")
    port = int(pick("REPORT_SMTP_PORT", "SMTP_PORT", "587") or "587")
    user = pick("REPORT_SMTP_USERNAME", "SMTP_USERNAME")
    password = pick("REPORT_SMTP_PASSWORD", "SMTP_PASSWORD")
    use_tls = (pick("REPORT_SMTP_USE_TLS", "SMTP_USE_TLS", "true") or "true").lower() == "true"
    use_ssl = (pick("REPORT_SMTP_USE_SSL", "SMTP_USE_SSL", "false") or "false").lower() == "true"
    sender = os.getenv("REPORT_FROM_EMAIL") or os.getenv("SMTP_FROM_EMAIL")
    reply_to = os.getenv("REPORT_REPLY_TO_EMAIL")  # optional

    if not host or not sender:
        raise RuntimeError(
            "SMTP host and a From: address must be set "
            "(REPORT_SMTP_HOST + REPORT_FROM_EMAIL, or fall back to SMTP_HOST + SMTP_FROM_EMAIL)"
        )

    return {
        "host": host, "port": port, "user": user, "password": password,
        "use_tls": use_tls, "use_ssl": use_ssl,
        "sender": sender, "reply_to": reply_to,
    }


def send_report(
    file_path: Path,
    from_date: str,
    to_date: str,
    row_count: int,
    recipients: list[str],
    dropped_team: list[str],
    healthy: bool,
) -> None:
    cfg = _resolve_smtp_config()
    subject, body = _build_subject_and_body(from_date, to_date, row_count, healthy)

    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = cfg["sender"]
    # Skip Reply-To when it would equal From — avoids redundant header.
    if cfg["reply_to"] and cfg["reply_to"].lower() != cfg["sender"].lower():
        msg["Reply-To"] = cfg["reply_to"]
    msg["To"] = ", ".join(recipients)
    msg.set_content(body)

    with file_path.open("rb") as f:
        msg.add_attachment(
            f.read(),
            maintype="application",
            subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=file_path.name,
        )

    if cfg["use_ssl"]:
        with smtplib.SMTP_SSL(cfg["host"], cfg["port"], timeout=30) as server:
            if cfg["user"] and cfg["password"]:
                server.login(cfg["user"], cfg["password"])
            server.send_message(msg)
    else:
        with smtplib.SMTP(cfg["host"], cfg["port"], timeout=30) as server:
            if cfg["use_tls"]:
                server.starttls()
            if cfg["user"] and cfg["password"]:
                server.login(cfg["user"], cfg["password"])
            server.send_message(msg)

    health_label = "healthy" if healthy else "degraded"
    print(f"Emailed {file_path.name} ({health_label}) to {', '.join(recipients)}")
    if dropped_team:
        print(f"Team distribution suppressed: {', '.join(dropped_team)}")


def main() -> int:
    args = parse_args()
    pg_conn_str = os.getenv("PG_CONN_STR")
    if not pg_conn_str:
        raise RuntimeError("PG_CONN_STR must be set")

    rows = fetch_rows(pg_conn_str, args.from_date, args.to_date)
    out_path = REPORTS_DIR / f"casino_daily_visits_{args.to_date}.xlsx"
    write_xlsx(rows, out_path)
    print(f"Wrote {len(rows)} rows -> {out_path}")

    if not args.no_email:
        if args.force_degraded:
            healthy = False
        else:
            healthy = pipeline_is_healthy(pg_conn_str, args.to_date)
        recipients, dropped = resolve_recipients(healthy)
        send_report(out_path, args.from_date, args.to_date, len(rows),
                    recipients, dropped, healthy)

    return 0


if __name__ == "__main__":
    sys.exit(main())
